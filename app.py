import io
import os
import uuid
import json
import time
import datetime as dt
from dataclasses import dataclass
from typing import List, Optional, Dict, Any, Tuple

import streamlit as st
import pandas as pd
import altair as alt

# 이미지 리사이즈/압축
try:
    from PIL import Image
    PIL_OK = True
except Exception:
    PIL_OK = False

# Supabase
try:
    from supabase import create_client
    SUPABASE_OK = True
except Exception:
    SUPABASE_OK = False

# 엑셀 생성
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# =========================
# 기본 설정
# =========================
APP_TITLE = "천안공장 HACCP 개선과제 시스템"
TABLE_TASKS = "haccp_tasks"
TABLE_PHOTOS = "haccp_task_photos"

STATUS_REGISTERED = "개선과제등록"
STATUS_PLANNED = "개선계획수립"
STATUS_DONE = "개선완료"

ALL_STATUSES = [STATUS_REGISTERED, STATUS_PLANNED, STATUS_DONE]

st.set_page_config(page_title=APP_TITLE, layout="wide")


# =========================
# 유틸
# =========================
def today_date() -> dt.date:
    return dt.date.today()

def parse_date_safe(x) -> Optional[dt.date]:
    if x is None or x == "":
        return None
    if isinstance(x, dt.date) and not isinstance(x, dt.datetime):
        return x
    if isinstance(x, dt.datetime):
        return x.date()
    try:
        return dt.date.fromisoformat(str(x)[:10])
    except Exception:
        return None

def iso_date(x: Optional[dt.date]) -> Optional[str]:
    if not x:
        return None
    return x.isoformat()

def now_iso() -> str:
    return dt.datetime.utcnow().isoformat()

def require_packages_or_stop():
    missing = []
    if not SUPABASE_OK:
        missing.append("supabase")
    if not PIL_OK:
        missing.append("Pillow(PIL)")
    if not OPENPYXL_OK:
        missing.append("openpyxl")
    if missing:
        st.error(
            "필수 라이브러리가 설치되지 않았습니다: "
            + ", ".join(missing)
            + "\n\nrequirements.txt에 추가 후 재배포하세요."
        )
        st.stop()

def require_secrets_or_stop(keys: List[str]):
    missing = [k for k in keys if k not in st.secrets or not str(st.secrets.get(k, "")).strip()]
    if missing:
        st.error("🚨 Secrets 누락: " + ", ".join(missing))
        st.stop()

def human_period_label(granularity: str) -> str:
    return "주간" if granularity == "weekly" else "월간"

def start_of_week(d: dt.date) -> dt.date:
    # 월요일 시작
    return d - dt.timedelta(days=d.weekday())

def month_start(d: dt.date) -> dt.date:
    return dt.date(d.year, d.month, 1)


# =========================
# Supabase 연결
# =========================
@st.cache_resource
def sb():
    require_secrets_or_stop(["SUPABASE_URL", "SUPABASE_ANON_KEY", "SUPABASE_SERVICE_KEY", "SUPABASE_BUCKET"])
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_SERVICE_KEY"]  # 안정적으로 CRUD하려면 서비스키 권장
    return create_client(url, key)

def bucket_name() -> str:
    return st.secrets["SUPABASE_BUCKET"]


# =========================
# 이미지 처리
# =========================
def compress_image(file_bytes: bytes, max_w: int = 1600, quality: int = 82) -> Tuple[bytes, str]:
    """
    - 입력: 원본 bytes
    - 출력: 압축된 JPEG bytes, 확장자("jpg")
    """
    if not PIL_OK:
        # Pillow 없으면 원본 반환 (그래도 동작은 하게)
        return file_bytes, "bin"

    img = Image.open(io.BytesIO(file_bytes))
    img = img.convert("RGB")

    w, h = img.size
    if w > max_w:
        new_h = int(h * (max_w / w))
        img = img.resize((max_w, new_h))

    out = io.BytesIO()
    img.save(out, format="JPEG", quality=quality, optimize=True)
    return out.getvalue(), "jpg"


# =========================
# DB/Storage 작업
# =========================
def fetch_tasks(limit: int = 2000) -> pd.DataFrame:
    client = sb()
    res = client.table(TABLE_TASKS).select("*").order("created_at", desc=True).limit(limit).execute()
    rows = res.data or []
    df = pd.DataFrame(rows)
    if df.empty:
        # 컬럼 기본 세팅
        df = pd.DataFrame(columns=[
            "id","created_at","issue_date","location","issue_text","reporter",
            "status","assignee","plan_due_date","action_text","action_date"
        ])
    # 날짜 변환
    for c in ["issue_date","plan_due_date","action_date"]:
        if c in df.columns:
            df[c] = df[c].apply(parse_date_safe)
    return df

def fetch_photos_for_task(task_id: str) -> pd.DataFrame:
    client = sb()
    res = client.table(TABLE_PHOTOS).select("*").eq("task_id", task_id).order("uploaded_at", desc=False).execute()
    rows = res.data or []
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["id","task_id","file_path","public_url","uploaded_at"])

def create_task(payload: Dict[str, Any]) -> Dict[str, Any]:
    client = sb()
    res = client.table(TABLE_TASKS).insert(payload).execute()
    if not res.data:
        raise RuntimeError("DB insert 실패")
    return res.data[0]

def update_task(task_id: str, patch: Dict[str, Any]) -> None:
    client = sb()
    client.table(TABLE_TASKS).update(patch).eq("id", task_id).execute()

def delete_task(task_id: str) -> None:
    # 사진 먼저 삭제
    photos = fetch_photos_for_task(task_id)
    for _, r in photos.iterrows():
        try:
            delete_photo(r["id"], r["file_path"])
        except Exception:
            pass
    client = sb()
    client.table(TABLE_TASKS).delete().eq("id", task_id).execute()

def upload_photo(task_id: str, original_name: str, file_bytes: bytes) -> Dict[str, Any]:
    client = sb()
    bname = bucket_name()

    # 압축/리사이즈
    compressed_bytes, ext = compress_image(file_bytes)

    safe_name = os.path.splitext(original_name)[0].replace(" ", "_")
    file_id = str(uuid.uuid4())
    path = f"{task_id}/{file_id}_{safe_name}.{ext}"

    # 업로드 (중복 방지: upsert=False)
    client.storage.from_(bname).upload(
        path=path,
        file=compressed_bytes,
        file_options={"content-type": "image/jpeg" if ext == "jpg" else "application/octet-stream"}
    )

    # public url
    public_url = client.storage.from_(bname).get_public_url(path)

    # DB 기록
    ins = {
        "task_id": task_id,
        "file_path": path,
        "public_url": public_url,
    }
    res = client.table(TABLE_PHOTOS).insert(ins).execute()
    if not res.data:
        raise RuntimeError("사진 메타 insert 실패")
    return res.data[0]

def delete_photo(photo_id: str, file_path: str) -> None:
    client = sb()
    bname = bucket_name()
    # storage 삭제
    try:
        client.storage.from_(bname).remove([file_path])
    except Exception:
        # storage 실패해도 메타 삭제는 진행(잔여파일은 나중에 정리 가능)
        pass
    # db 삭제
    client.table(TABLE_PHOTOS).delete().eq("id", photo_id).execute()

def replace_photo(photo_id: str, old_path: str, task_id: str, original_name: str, file_bytes: bytes) -> None:
    # 새 업로드
    new_meta = upload_photo(task_id, original_name, file_bytes)
    # 기존 삭제(새 업로드 성공 후)
    delete_photo(photo_id, old_path)
    # (선택) 새 사진을 "대표"로 만들고 싶으면 여기서 정렬/플래그 로직 추가 가능


# =========================
# 리포트/엑셀
# =========================
def filter_by_period(df: pd.DataFrame, start: dt.date, end: dt.date) -> pd.DataFrame:
    if df.empty:
        return df
    # issue_date 기준
    m = df["issue_date"].apply(lambda d: d is not None and start <= d <= end)
    return df[m].copy()

def summarize(df: pd.DataFrame) -> Dict[str, Any]:
    total = len(df)
    done = int((df["status"] == STATUS_DONE).sum()) if not df.empty else 0
    by_loc = (df.groupby("location")["id"].count().sort_values(ascending=False).reset_index(name="count")
              if not df.empty else pd.DataFrame(columns=["location","count"]))
    by_status = (df.groupby("status")["id"].count().reindex(ALL_STATUSES).fillna(0).reset_index(name="count")
                 if not df.empty else pd.DataFrame({"status": ALL_STATUSES, "count":[0,0,0]}))
    return {"total": total, "done": done, "by_loc": by_loc, "by_status": by_status}

def build_timeseries(df: pd.DataFrame, granularity: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["period","발굴","완료"])

    rows = []
    for _, r in df.iterrows():
        d = r["issue_date"]
        if not d:
            continue
        if granularity == "weekly":
            p = start_of_week(d)
        else:
            p = month_start(d)
        rows.append((p, 1, 1 if r["status"] == STATUS_DONE else 0))

    ts = pd.DataFrame(rows, columns=["period","발굴","완료"])
    if ts.empty:
        return pd.DataFrame(columns=["period","발굴","완료"])
    ts = ts.groupby("period")[["발굴","완료"]].sum().reset_index()
    ts["period_str"] = ts["period"].astype(str)
    return ts.sort_values("period")

def export_excel_links(tasks_df: pd.DataFrame) -> bytes:
    """
    기본: 사진은 '하이퍼링크'로 제공 (가장 안정적)
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl 미설치")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리포트"

    headers = [
        "발굴일", "공정/장소", "개선 필요사항", "발견자",
        "진행상태", "담당자", "개선계획일", "개선내용", "개선완료일", "사진(링크)"
    ]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # task별 사진 링크 1줄로 합치기
    client = sb()

    for i, r in tasks_df.reset_index(drop=True).iterrows():
        task_id = r["id"]
        photos = client.table(TABLE_PHOTOS).select("public_url").eq("task_id", task_id).execute().data or []
        links = [p["public_url"] for p in photos if p.get("public_url")]
        link_text = " | ".join(links) if links else ""

        ws.append([
            iso_date(r.get("issue_date")),
            r.get("location",""),
            r.get("issue_text",""),
            r.get("reporter",""),
            r.get("status",""),
            r.get("assignee",""),
            iso_date(r.get("plan_due_date")),
            r.get("action_text",""),
            iso_date(r.get("action_date")),
            link_text
        ])

    # 열 너비/줄바꿈
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 55

    for row in range(2, ws.max_row+1):
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=10).alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# UI 구성
# =========================
require_packages_or_stop()

st.title(APP_TITLE)

# 상단 안내/상태
with st.expander("✅ 운영 체크(필수 설정 확인)", expanded=False):
    st.write("아래 항목이 모두 OK면 앱은 안정적으로 동작합니다.")
    ok = True
    for k in ["SUPABASE_URL","SUPABASE_ANON_KEY","SUPABASE_SERVICE_KEY","SUPABASE_BUCKET"]:
        v = str(st.secrets.get(k, "")).strip()
        st.write(f"- {k}: {'OK' if v else '❌ 누락'}")
        ok = ok and bool(v)
    if not ok:
        st.warning("Secrets를 먼저 채워주세요. (App settings → Secrets)")
        st.stop()

# 데이터 로드
df_all = fetch_tasks()

tabs = st.tabs([
    "대시보드/보고서",
    "개선과제등록",
    "개선계획수립",
    "개선완료 입력",
    "조회/관리"
])

# =========================
# 1) 대시보드/보고서
# =========================
with tabs[0]:
    st.subheader("대시보드/보고서")

    colA, colB, colC = st.columns([1,1,2])
    with colA:
        granularity = st.radio("집계 단위", ["weekly","monthly"], format_func=human_period_label, horizontal=True)
    with colB:
        start = st.date_input("시작일", value=today_date() - dt.timedelta(days=30))
    with colC:
        end = st.date_input("종료일", value=today_date())

    df = filter_by_period(df_all, start, end)
    s = summarize(df)

    m1, m2, m3 = st.columns(3)
    m1.metric("총 발굴건수", s["total"])
    m2.metric("개선완료 건수", s["done"])
    m3.metric("완료율", f"{(s['done']/s['total']*100):.1f}%" if s["total"] else "0.0%")

    ts = build_timeseries(df, granularity)
    if not ts.empty:
        chart = alt.Chart(ts).transform_fold(
            ["발굴","완료"], as_=["구분","건수"]
        ).mark_line(point=True).encode(
            x=alt.X("period_str:N", title="기간"),
            y=alt.Y("건수:Q", title="건수"),
            color="구분:N"
        )
        st.altair_chart(chart, use_container_width=True)
    else:
        st.info("선택한 기간에 데이터가 없습니다.")

    c1, c2 = st.columns(2)
    with c1:
        st.write("공정(실)별 발굴 건수")
        if not s["by_loc"].empty:
            st.dataframe(s["by_loc"], use_container_width=True, hide_index=True)
        else:
            st.write("-")

    with c2:
        st.write("진행상태별 건수")
        if not s["by_status"].empty:
            st.dataframe(s["by_status"], use_container_width=True, hide_index=True)
        else:
            st.write("-")

    st.divider()
    st.subheader("보고서/엑셀 출력")

    if st.button("📄 엑셀 다운로드 생성(사진 링크 포함)", type="primary"):
        if df.empty:
            st.warning("다운로드할 데이터가 없습니다.")
        else:
            try:
                xlsx_bytes = export_excel_links(df)
                st.download_button(
                    "⬇️ HACCP_리포트.xlsx 다운로드",
                    data=xlsx_bytes,
                    file_name=f"HACCP_리포트_{start}_{end}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("엑셀 생성 완료!")
            except Exception as e:
                st.error(f"엑셀 생성 실패: {e}")


# =========================
# 2) 개선과제등록
# =========================
with tabs[1]:
    st.subheader("개선과제등록 (발굴/등록)")

    with st.form("form_register", clear_on_submit=True):
        c1, c2, c3 = st.columns([1,1,2])
        issue_date = c1.date_input("일시", value=today_date())
        location = c2.text_input("공정/장소", placeholder="예: 전처리실")
        reporter = c3.text_input("발견자", placeholder="예: 품질보증팀")

        issue_text = st.text_area("개선 필요사항", height=120, placeholder="무엇이 문제인지 구체적으로 작성")

        st.caption("사진은 여러 장 업로드 가능 (자동 리사이즈/압축)")
        photos = st.file_uploader(
            "사진 업로드",
            type=["jpg","jpeg","png","webp"],
            accept_multiple_files=True
        )

        submitted = st.form_submit_button("✅ 등록하기", type="primary")

    if submitted:
        if not location.strip():
            st.warning("공정/장소를 입력하세요.")
            st.stop()
        if not reporter.strip():
            st.warning("발견자를 입력하세요.")
            st.stop()
        if not issue_text.strip():
            st.warning("개선 필요사항을 입력하세요.")
            st.stop()

        try:
            payload = {
                "issue_date": iso_date(issue_date),
                "location": location.strip(),
                "issue_text": issue_text.strip(),
                "reporter": reporter.strip(),
                "status": STATUS_REGISTERED,
                "assignee": None,
                "plan_due_date": None,
                "action_text": None,
                "action_date": None,
            }
            new_task = create_task(payload)
            task_id = new_task["id"]

            # 사진 업로드
            uploaded = 0
            if photos:
                for f in photos:
                    upload_photo(task_id, f.name, f.getvalue())
                    uploaded += 1

            st.success(f"등록 완료! (사진 {uploaded}장 업로드)")
            st.rerun()
        except Exception as e:
            st.error(f"등록 실패: {e}")


# =========================
# 3) 개선계획수립
# =========================
with tabs[2]:
    st.subheader("개선계획수립 (담당자/일정 지정)")

    df_plan = df_all[df_all["status"].isin([STATUS_REGISTERED, STATUS_PLANNED])].copy()
    if df_plan.empty:
        st.info("계획수립 대상이 없습니다.")
    else:
        df_plan["표시"] = df_plan.apply(
            lambda r: f"[{r['status']}] {r.get('issue_date')} / {r.get('location','')} / {str(r.get('issue_text',''))[:40]}",
            axis=1
        )
        pick = st.selectbox("대상 선택", df_plan["표시"].tolist())
        row = df_plan[df_plan["표시"] == pick].iloc[0].to_dict()
        task_id = row["id"]

        st.write("**개선 필요사항**")
        st.write(row.get("issue_text",""))

        c1, c2, c3 = st.columns([1,1,2])
        assignee = c1.text_input("담당자(팀/부서)", value=row.get("assignee") or "")
        plan_due = c2.date_input("개선계획(일정)", value=row.get("plan_due_date") or today_date())
        status_now = c3.selectbox("진행상태", [STATUS_REGISTERED, STATUS_PLANNED], index=1 if row["status"]==STATUS_PLANNED else 0)

        if st.button("💾 계획 저장", type="primary"):
            try:
                update_task(task_id, {
                    "assignee": assignee.strip() if assignee.strip() else None,
                    "plan_due_date": iso_date(plan_due),
                    "status": status_now
                })
                st.success("저장 완료")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")

        st.divider()
        st.write("### 사진")
        photos_df = fetch_photos_for_task(task_id)
        if photos_df.empty:
            st.caption("등록된 사진이 없습니다.")
        else:
            # 그리드 출력
            cols = st.columns(3)
            for i, r in photos_df.iterrows():
                with cols[i % 3]:
                    st.image(r["public_url"], use_container_width=True)
                    cdel, crep = st.columns(2)
                    with cdel:
                        if st.button("삭제", key=f"del_{r['id']}"):
                            try:
                                delete_photo(r["id"], r["file_path"])
                                st.success("삭제 완료")
                                st.rerun()
                            except Exception as e:
                                st.error(f"삭제 실패: {e}")
                    with crep:
                        newf = st.file_uploader("교체", type=["jpg","jpeg","png","webp"], key=f"rep_{r['id']}")
                        if newf is not None:
                            if st.button("교체 적용", key=f"repbtn_{r['id']}"):
                                try:
                                    replace_photo(r["id"], r["file_path"], task_id, newf.name, newf.getvalue())
                                    st.success("교체 완료")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"교체 실패: {e}")

        st.divider()
        st.write("### 사진 추가 업로드")
        add_files = st.file_uploader("추가 사진", type=["jpg","jpeg","png","webp"], accept_multiple_files=True, key="plan_add_photos")
        if st.button("➕ 추가 업로드"):
            if not add_files:
                st.warning("추가할 사진을 선택하세요.")
            else:
                try:
                    for f in add_files:
                        upload_photo(task_id, f.name, f.getvalue())
                    st.success("추가 업로드 완료")
                    st.rerun()
                except Exception as e:
                    st.error(f"업로드 실패: {e}")


# =========================
# 4) 개선완료 입력
# =========================
with tabs[3]:
    st.subheader("개선완료 입력 (조치내용/완료일)")

    df_done = df_all.copy()
    if df_done.empty:
        st.info("데이터가 없습니다.")
    else:
        df_done["표시"] = df_done.apply(
            lambda r: f"[{r['status']}] {r.get('issue_date')} / {r.get('location','')} / {str(r.get('issue_text',''))[:40]}",
            axis=1
        )
        pick = st.selectbox("대상 선택", df_done["표시"].tolist())
        row = df_done[df_done["표시"] == pick].iloc[0].to_dict()
        task_id = row["id"]

        st.write("**개선 필요사항**")
        st.write(row.get("issue_text",""))

        c1, c2 = st.columns([2,1])
        action_text = c1.text_area("개선내용(조치내용)", value=row.get("action_text") or "", height=140)
        action_date = c2.date_input("개선완료일", value=row.get("action_date") or today_date())
        status_new = st.selectbox("진행상태", ALL_STATUSES, index=ALL_STATUSES.index(row.get("status") or STATUS_REGISTERED))

        # 완료 버튼은 조치내용 없으면 막기(오류 최소)
        if st.button("✅ 완료 저장", type="primary"):
            if status_new == STATUS_DONE and not action_text.strip():
                st.warning("완료 처리하려면 '개선내용'을 입력하세요.")
                st.stop()
            try:
                update_task(task_id, {
                    "action_text": action_text.strip() if action_text.strip() else None,
                    "action_date": iso_date(action_date) if status_new == STATUS_DONE else iso_date(action_date),
                    "status": status_new
                })
                st.success("저장 완료")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")

        st.divider()
        st.write("### 사진(즉시 확인)")
        photos_df = fetch_photos_for_task(task_id)
        if photos_df.empty:
            st.caption("등록된 사진이 없습니다.")
        else:
            cols = st.columns(3)
            for i, r in photos_df.iterrows():
                with cols[i % 3]:
                    st.image(r["public_url"], use_container_width=True)


# =========================
# 5) 조회/관리
# =========================
with tabs[4]:
    st.subheader("조회/관리")

    c1, c2, c3, c4 = st.columns([1,1,1,2])
    with c1:
        f_status = st.multiselect("상태", ALL_STATUSES, default=ALL_STATUSES)
    with c2:
        f_loc = st.text_input("공정/장소 필터", placeholder="예: 전처리")
    with c3:
        f_reporter = st.text_input("발견자 필터", placeholder="예: 품질")
    with c4:
        kw = st.text_input("키워드 검색", placeholder="개선 필요사항/개선내용 검색")

    df = df_all.copy()
    if f_status:
        df = df[df["status"].isin(f_status)]
    if f_loc.strip():
        df = df[df["location"].fillna("").str.contains(f_loc.strip(), na=False)]
    if f_reporter.strip():
        df = df[df["reporter"].fillna("").str.contains(f_reporter.strip(), na=False)]
    if kw.strip():
        k = kw.strip()
        df = df[
            df["issue_text"].fillna("").str.contains(k, na=False) |
            df["action_text"].fillna("").str.contains(k, na=False)
        ]

    show_cols = [
        "issue_date","location","issue_text","reporter","status",
        "assignee","plan_due_date","action_text","action_date","id"
    ]
    st.dataframe(df[show_cols], use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("삭제(주의)")

    if df.empty:
        st.caption("삭제할 항목이 없습니다.")
    else:
        df["표시"] = df.apply(
            lambda r: f"{r.get('issue_date')} / {r.get('location','')} / {str(r.get('issue_text',''))[:40]}",
            axis=1
        )
        pick = st.selectbox("삭제할 항목 선택", df["표시"].tolist(), key="del_pick")
        row = df[df["표시"] == pick].iloc[0].to_dict()
        task_id = row["id"]

        st.warning("삭제하면 해당 항목 + 연결된 사진이 모두 삭제됩니다.")
        confirm = st.checkbox("정말 삭제할게요(체크 필요)")
        if st.button("🗑 삭제 실행", disabled=not confirm):
            try:
                delete_task(task_id)
                st.success("삭제 완료")
                st.rerun()
            except Exception as e:
                st.error(f"삭제 실패: {e}")

# 끝
